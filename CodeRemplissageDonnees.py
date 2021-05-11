import openpyxl as xl
from random import *

wb = xl.load_workbook('Donnees1.xlsx')
sheet = wb['Feuil1']

for row in range(2,21):
    cell = sheet.cell(row,2)
    cell.value = 1800*random()
    
for row in range(21,30):
    cell = sheet.cell(row,2)
    cell.value = randint(700,900) + random()
    
for row in range(30,sheet.max_row +1):
    cell = sheet.cell(row,2)
    cell.value = 770
    
    
for row in range(2,21):
    cell = sheet.cell(row,3)
    cell.value = 10*random()
    
for row in range(21,30):
    cell = sheet.cell(row,3)
    cell.value = randint(2,5) + random()
    
for row in range(30,sheet.max_row +1):
    cell = sheet.cell(row,3)
    cell.value = randint(0,2) + random()
    
for row in range(2,21):
    cell = sheet.cell(row,4)
    cell.value = 12*random()
    
for row in range(21,30):
    cell = sheet.cell(row,4)
    cell.value = randint(3,6) + random()
    
for row in range(30,sheet.max_row +1):
    cell = sheet.cell(row,4)
    cell.value = 4.2
    
for row in range(2,21):
    cell = sheet.cell(row,5)
    cell.value = 16*random()
    
for row in range(21,30):
    cell = sheet.cell(row,5)
    cell.value = randint(3,8) + random()
    
for row in range(30,sheet.max_row +1):
    cell = sheet.cell(row,5)
    cell.value = randint(0,4) + random()

for row in range(2,21):
    cell = sheet.cell(row,6)
    cell.value = 12*random()
    
for row in range(21,30):
    cell = sheet.cell(row,6)
    cell.value = randint(2,6) + random()
    
for row in range(30,sheet.max_row +1):
    cell = sheet.cell(row,6)
    cell.value = 4.5
    
for row in range(2,21):
    cell = sheet.cell(row,7)
    cell.value = 480*random()
    
for row in range(21,30):
    cell = sheet.cell(row,7)
    cell.value = randint(120,240) + random()
    
for row in range(30,sheet.max_row +1):
    cell = sheet.cell(row,7)
    cell.value = 170
    
for row in range(2,21):
    cell = sheet.cell(row,8)
    cell.value = 12*random()
    
for row in range(21,30):
    cell = sheet.cell(row,8)
    cell.value = randint(3,6) + random()
    
for row in range(30,sheet.max_row +1):
    cell = sheet.cell(row,8)
    cell.value = 4
    
for row in range(2,21):
    cell = sheet.cell(row,9)
    cell.value = 30*random()
    
for row in range(21,30):
    cell = sheet.cell(row,9)
    cell.value = randint(7,15) + random()
    
for row in range(30,sheet.max_row +1):
    cell = sheet.cell(row,9)
    cell.value = randint(9,11) + random()
    
for row in range(2,21):
    cell = sheet.cell(row,10)
    cell.value = 68*random()
    
for row in range(21,30):
    cell = sheet.cell(row,10)
    cell.value = randint(23,34) + random()
    
for row in range(30,sheet.max_row +1):
    cell = sheet.cell(row,10)
    cell.value = 25
    
for row in range(2,21):
    cell = sheet.cell(row,11)
    cell.value = 2*random()
    
for row in range(21,30):
    cell = sheet.cell(row,11)
    cell.value = random()
    
for row in range(30,sheet.max_row +1):
    cell = sheet.cell(row,11)
    cell.value = 0.45
    
for row in range(2,21):
    cell = sheet.cell(row,12)
    cell.value = 2*random()
    
for row in range(21,30):
    cell = sheet.cell(row,12)
    cell.value = random()
    
for row in range(30,sheet.max_row +1):
    cell = sheet.cell(row,12)
    cell.value = 0.6
    
for row in range(2,21):
    cell = sheet.cell(row,13)
    cell.value = 28*random()
    
for row in range(21,30):
    cell = sheet.cell(row,13)
    cell.value = randint(11,14) + random()

for row in range(30,sheet.max_row + 1):
    cell = sheet.cell(row,13)
    cell.value = 0

for row in range(2,21):
    cell = sheet.cell(row,14)
    cell.value = 20*random()
    
for row in range(21,30):
    cell = sheet.cell(row,14)
    cell.value = 5
    
for row in range(30,sheet.max_row +1):
    cell = sheet.cell(row,14)
    cell.value = randint(5,9) + random()
    
for row in range(2,21):
    cell = sheet.cell(row,15)
    cell.value = 20*random()
    
for row in range(21,30):
    cell = sheet.cell(row,15)
    cell.value = 9
    
for row in range(30,sheet.max_row +1):
    cell = sheet.cell(row,15)
    cell.value = 10
    
for row in range(2,21):
    cell = sheet.cell(row,16)
    cell.value = 6*random()
    
for row in range(21,sheet.max_row +1):
    cell = sheet.cell(row,16)
    cell.value = 3.5
    
for row in range(2,21):
    cell = sheet.cell(row,17)
    cell.value = 300*random()
    
for row in range(21,30):
    cell = sheet.cell(row,17)
    cell.value = 150
    
for row in range(30,sheet.max_row +1):
    cell = sheet.cell(row,17)
    cell.value = randint(100,149) + random()
    
for row in range(2,21):
    cell = sheet.cell(row,18)
    cell.value = 2000*random()
    
for row in range(21,sheet.max_row +1):
    cell = sheet.cell(row,18)
    cell.value = randint(1000,1300) + random()
    
for row in range(2,21):
    cell = sheet.cell(row,19)
    cell.value = 6*random()
    
for row in range(21,sheet.max_row +1):
    cell = sheet.cell(row,19)
    cell.value = 1.5 + random()
    
for row in range(2,21):
    cell = sheet.cell(row,20)
    cell.value = 800*random()
    
for row in range(21,30):
    cell = sheet.cell(row,20)
    cell.value = randint(320,420) + random()

for row in range(30,sheet.max_row+1):
    cell = sheet.cell(row,20)
    cell.value = 400
    
for row in range(2,21):
    cell = sheet.cell(row,21)
    cell.value = 10*random()
    
for row in range(21,30):
    cell = sheet.cell(row,21)
    cell.value = randint(2,5) + random()
    
for row in range(30,sheet.max_row +1):
    cell = sheet.cell(row,21)
    cell.value = 4 + random()

for row in range(2,21):
    cell = sheet.cell(row,22)
    cell.value = 1200*random()
    
for row in range(21,sheet.max_row +1):
    cell = sheet.cell(row,22)
    cell.value = 700

for row in range(2,21):
    cell = sheet.cell(row,23)
    cell.value = 100*random()
    
for row in range(21,30):
    cell = sheet.cell(row,23)
    cell.value = 55
    
for row in range(30,sheet.max_row +1):
    cell = sheet.cell(row,23)
    cell.value = randint(60,69) + random()
    
for row in range(2,21):
    cell = sheet.cell(row,24)
    cell.value = 22*random()
    
for row in range(21,30):
    cell = sheet.cell(row,24)
    cell.value = randint(8,10) + random()
    
for row in range(30,sheet.max_row +1):
    cell = sheet.cell(row,24)
    cell.value = 11

wb.save('Donnees.xlsx')